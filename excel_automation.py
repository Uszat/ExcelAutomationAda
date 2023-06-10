import openpyxl
from openpyxl import Workbook
import os
import time
import sys

#defines
OFFSET_TO_START_FROM_ONE = 1 #starts from 0 but I want to start from 1
OFFSET_TO_CALC_NO_PPL = 1 #people in sheet actually start from 2nd pos
MIN_DAYS_TO_MATCH = 1 #minimum number of days that have to match to pair up people
MIN_SPORTS_TO_MATCH = 1 #minimum number of sports that have to match to pair up people
NUMBER_OF_UNMATCHED_PEOPLE = 2


#specify the files location (or path)
cwd = os.path.dirname(os.path.abspath(__file__))
file = cwd + '\\fitness_buddy_ada.xlsx'

#create an empty list to append values later on
values = []

#index of pairs with their buddy
pairIndex = 0

noOfSports = 0

#global index of how many people were put into a sheet
numberOfPeopleInSheet = 0

#list of groups of people who look for a pair
listOfGroups = []
#list of pairs of people who already have a pair
hasPairGroup = []
#list of names that has already been assigned with their specified pairName
namesAlreadyTakenCareOf = []
#list of names that are matched by thier sport preference
matchedBySport = []
#list of objects that are matched by thier sport preference
objectsMatchedBySport = []
#list of leftover people that didnt get a pair cause they were only ones wanting certain sport or just number of all people is odd
notMatchedBySport = []

#open file and worksheet
workbook = openpyxl.load_workbook(file)
worksheet = workbook.active
noOfEntries = worksheet.max_row - OFFSET_TO_CALC_NO_PPL

#create new workbook for assigned pairs
wbAssigned = Workbook()
wsAssigned =  wbAssigned.active

nie = 'Nie - dołączam z własnym Fitness Buddy'

#create a class of people
class Person(object):
    def __init__(self, number):
        self.number = number
        self.sportSplit = ''

    #overload equals operator
    def __eq__(self, other):
        if (self.wantToBeMatched == other.wantToBeMatched and 
            self.genderBucket == other.genderBucket and
            self.freqz == other.freqz and
            self.onlineStationSet == other.onlineStationSet
            # and self.town == other.town #uncomment if want to add town constraint 
            ):
            return True
        else:
            return False

    #print person data
    def showData(self):
        print("number \t\t",        self.number + OFFSET_TO_START_FROM_ONE)
        print("name \t\t",          self.name)
        print("wantToBeMatched \t", self.wantToBeMatched)
        print("nameToPair \t",      self.nameToPair)
        print("discipline \t",      self.discipline)
        print(" ")


#fill People's attributes with values from the worksheet
def initObject():
    for index, person in enumerate(people):
        i = index + 2 #index starts from zero so make i from 2 as the real rows start from 2nd
        person.name =           worksheet['B' + str(i)].value
        person.wantToBeMatched =  worksheet['D' + str(i)].value
        person.nameToPair =     worksheet['E' + str(i)].value
        person.discipline =     worksheet['F' + str(i)].value


def openFile():
    try:
        os.startfile(cwd + '\\fitness_buddy_ada_assigned.xlsx')
    except:
        print("error opening Excel")


def forceCloseFile():
    try:
        os.system('TASKKILL /F /IM EXCEL.EXE')
        print("No worries, it's just a warning that Excel isn't open\n")
        time.sleep(1)
    except:
        print("error killing Excel")


def putPeopleWithPairToGroup():
    global hasPairGroup
    global namesAlreadyTakenCareOf
    
    #add all people that want a pair to a hasPairGroup
    for person in people:
        if(person.wantToBeMatched == nie and person.nameToPair is not None):
            hasPairGroup.append([person])

    #to every person in hasPairGroup add their matching pairName and add them as new element in sublist
    for idx, groupElement in enumerate(hasPairGroup):
        for person in people:
            if(person.name == groupElement[idx].nameToPair):
                hasPairGroup[idx].extend([person])

    #make an array containing the names we don't want to match in the future as they are already done
    for groupElement in hasPairGroup:
        for singleElement in groupElement:
            namesAlreadyTakenCareOf.append(singleElement.name)


#changes the main list "people" by removing already used people (the ones who already have been matched by pair)
def removeUsedPersonFromList():
    global people
    updated_people = []
    for person in people:
        if person.name not in namesAlreadyTakenCareOf:
            updated_people.append(person) 

    people = updated_people
    # for person in people:
    #     print(person.name)


# match people from the list "people" by one shared sport category
def matchPeopleBySport(iter):
    sportLists = [[], [], []]  # List with three dimensions for each sport type
    sport_type = ['Koszykówka', 'Siatkówka', 'Yoga']  # Map index to sport type
    global noOfSports 
    global matchedBySport
    global notMatchedBySport

    noOfSports = len(sport_type)
    matchedBySport = []
    notMatchedBySport = []

    for person in people:
        try:
            person.sportSplit = person.discipline.split(sep=",")  # separate every comma and put values into a list [Kosz, siatka, yoga] -> [[kosz], [siatka], [yoga]]
            person.sportSplit = [item.strip() for item in person.sportSplit] #strip it from whitespaces
        except AttributeError:
            print("ERROR")
            print("Program thinks there are more people in Excel than it actually is. Try copying the content of the excel to another workbook and save it anew")
            print("TERMINATING")
            sys.exit()

        
        for sport in sport_type:
            if sport in person.sportSplit:
                if(sport == 'Koszykówka'):
                    sportLists[(0+iter)%3].append(person.name)
                if(sport == 'Siatkówka'):
                    sportLists[(1+iter)%3].append(person.name)
                if(sport == 'Yoga'):
                    sportLists[(2+iter)%3].append(person.name)
    
    # The idea here is to go through the list containing 3 sublists with people, one sublist for every sport and inside a person that wanted that sport
    # Same person can be in a few sublists
    # Then if only one person chose certain sport - remove that person, we will not match them either way. Add it to unmatched group
    # Otherwise if the sublist is of length 2 or more then
    #   Check if its length is even
    #       if so just connect pairs of people together
    #   if not even
    #       connect them in pairs till one person is left and then add it to unmatched group
    # In both cases (even or odd) every time two people are added remove all instances of those names from sportLists so that they are not paired again

    while len(sportLists) > 0:
        for sportGroup in sportLists:
            if len(sportGroup) == 1:
                notMatchedBySport.append(sportGroup[0]) #sportgroup in that case is always length 1 so take the first element
                sportLists.remove(sportGroup) #cannot pair one person so just remove that one-person group [[MP], [MP, KJ]] -> [[MP, KJ]]

            elif len(sportGroup) >= 2: #there are at least two people to pair them
                if len(sportGroup) % 2 == 0: #even groups only
                    while len(sportGroup) > 0: #add people to matchedbysport and remove from sportGroup and sportLists till none left in sportGroup
                        matchedBySport.append(sportGroup[:2])  # Add first two people from the sub list to matchedBySport list
                        
                        ### removing used people from sportLists ###
                        for name in sportGroup[:2]:
                            for group in sportLists: #[[MP, KJ, WB, DZ], [MP, XD, DZ]] tak wylada sportlists, wiec biore jej sub listy i wrzucam do group
                                while name in group: #teraz biore serio imiona z sub list - group: [MP, KJ, WB, DZ] 
                                    group.remove(name) #here remove people from the sportLists group [[MP, KJ, WB, DZ], [MP, XD, DZ]] -> [[WB, DZ], [XD, DZ]]
                        ### removing used people from sportLists ###
                        
                else:  # Odd number of people
                    while len(sportGroup) > 1:
                        matchedBySport.append(sportGroup[:2])  # Add first two people from the sub list to matchedBySport list 

                        ### removing used people from sportLists ###
                        for name in sportGroup[:2]:
                            for group in sportLists: #[[MP, KJ, WB, DZ], [MP, XD, DZ]] tak wylada sportlists, wiec biore jej sub listy i wrzucam do group
                                while name in group: #teraz biore serio imiona z sub list
                                    group.remove(name) #here remove people from the sportLists group
                        ### removing used people from sportLists ###

                    # Doubles the wrong result notMatchedBySport.append(sportGroup[0]) #sportgroup in that case is always length 1 so take the first element
                        
            sportLists = [sublist for sublist in sportLists if sublist != []] #subLists looks like this [[], [], ['Mateusz Krzak']] and the length of 
            # such list is 3, it counts [] as 1 list, so I remove empty lists with it and in the end I get this [['Mateusz Krzak']]    

    return len(notMatchedBySport)


# matchedBySport list only operates on name and surname [['Mateusz Krzak']] instead of objects (which contain all data about a person)
# So find matching name between list matchedBySport and person.name and add this object person to the list objectMatchedBySport
def fillObjectList():

    newPair = True
    index = 0
    for group in matchedBySport:
        for name in group:
            for person in people:
                if person.name == name:
                    if newPair == True:
                        objectsMatchedBySport.append([person]) # appending to add a new index of the list
                        newPair = False
                    else:
                        objectsMatchedBySport[index].extend([person]) #extend to add another item to the list in certain index (so that the list consists of sublists with pairs). idx/2 cause you extend every second person and int() is just to take whole value
                        newPair = True
                        index += 1

#inserting people's all data into cells
def putPeopleInCell():
    global pairIndex
    global numberOfPeopleInSheet    
    for id in range(len(objectsMatchedBySport)):
        for idd in range(len(objectsMatchedBySport[id])):
            pairIndex+=1
            numberOfPeopleInSheet+=1
            wsAssigned['A' + str(pairIndex)] = numberOfPeopleInSheet
            wsAssigned['B' + str(pairIndex)] = objectsMatchedBySport[id][idd].name
            wsAssigned['C' + str(pairIndex)] = objectsMatchedBySport[id][idd].wantToBeMatched 
            wsAssigned['D' + str(pairIndex)] = objectsMatchedBySport[id][idd].nameToPair 
            wsAssigned['E' + str(pairIndex)] = objectsMatchedBySport[id][idd].discipline

        pairIndex+=1


#inserting people's all data into cells
def putHasBuddyPeopleInCell():
    global pairIndex
    global numberOfPeopleInSheet    
    for id in range(len(hasPairGroup)):
        for idd in range(len(hasPairGroup[id])):
            pairIndex+=1
            numberOfPeopleInSheet+=1
            wsAssigned['A' + str(pairIndex)] = numberOfPeopleInSheet
            wsAssigned['B' + str(pairIndex)] = hasPairGroup[id][idd].name
            wsAssigned['C' + str(pairIndex)] = hasPairGroup[id][idd].wantToBeMatched 
            wsAssigned['D' + str(pairIndex)] = hasPairGroup[id][idd].nameToPair 
            wsAssigned['E' + str(pairIndex)] = hasPairGroup[id][idd].discipline

        pairIndex+=1


#show all People's data
def showAllPeople():
    for person in people:
        person.showData()


############ MAIN ################
if __name__ == "__main__":
    
    print("Starting...")

    #create list of objects People
    people = []
    for i in range(noOfEntries):
        people.append(Person(i))  

    initObject()

    #leave this commented
    #showAllPeople()

    # force shut excel file to be able to amend the assigned excel
    forceCloseFile()

    putPeopleWithPairToGroup()
    removeUsedPersonFromList()
    iter = 0
    while matchPeopleBySport(iter) >= NUMBER_OF_UNMATCHED_PEOPLE and iter < noOfSports: #do as many shuffles as there are sports to find the best possible assignment
        iter+=1
    fillObjectList()

    putPeopleInCell()
    putHasBuddyPeopleInCell()

    #save and open new worksheet with added pairs
    wbAssigned.save(cwd + '\\fitness_buddy_ada_assigned.xlsx')
    openFile()

    #check if all people were successfully put into sheet
    if(noOfEntries is not numberOfPeopleInSheet):
        print("MISSING PEOPLE!")
        print(noOfEntries - numberOfPeopleInSheet, "people were ommitted")
    else:
        print("Success!")